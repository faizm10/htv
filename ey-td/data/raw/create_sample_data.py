#!/usr/bin/env python3
"""
Create sample Excel files for testing the EY data pipeline
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta

def create_customer_data():
    """Create sample customer data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = ['CustomerID', 'LegalID', 'LegalExpiredDate', 'Country', 'City']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    data = [
        ['CUST-001', 'AB123456', '2023-12-31', 'US', 'New York'],
        ['CUST-002', 'CD789012', '2024-06-15', 'CA', 'Toronto'],
        ['CUST-003', 'EF345678', '2022-08-20', 'UK', 'London'],
        ['CUST-004', 'AB123456', '2024-03-10', 'US', 'Chicago'],  # Duplicate legal ID
        ['CUST-005', 'GH901234', '2025-01-15', 'AU', 'Sydney']
    ]
    
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save('Bank1_Mock_Customer.xlsx')
    print("Created Bank1_Mock_Customer.xlsx")

def create_accounts_cursav():
    """Create sample current/savings accounts"""
    wb = Workbook()
    ws = wb.active
    ws.title = "CurrentSavings"
    
    # Headers
    headers = ['AccountID', 'AccountType', 'AccountStatus', 'UnclearedBalance', 'LockedAmount']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    data = [
        ['ACC-001', 'Savings', 'Active', 15000.00, 0.00],
        ['ACC-002', 'Checking', 'Dormant', 7500.00, 2000.00],  # Dormant high balance
        ['ACC-003', 'Savings', 'Active', 3000.00, 0.00],
        ['ACC-004', 'Checking', 'Dormant', 12000.00, 4000.00],  # Dormant high balance, high locked ratio
        ['ACC-005', 'Savings', 'Active', 800.00, 0.00]
    ]
    
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save('Bank1_Mock_CurSav_Accounts.xlsx')
    print("Created Bank1_Mock_CurSav_Accounts.xlsx")

def create_accounts_fixed():
    """Create sample fixed term accounts"""
    wb = Workbook()
    ws = wb.active
    ws.title = "FixedTerm"
    
    # Headers
    headers = ['AccountID', 'AccountType', 'AccountStatus', 'MaturityDate']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    data = [
        ['FIX-001', 'Term Deposit', 'Active', '2023-12-01'],  # Past maturity
        ['FIX-002', 'CD Account', 'Active', '2024-01-15'],    # Past maturity
        ['FIX-003', 'Term Deposit', 'Closed', '2022-06-30'],
        ['FIX-004', 'CD Account', 'Active', '2025-03-15']
    ]
    
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save('Bank1_Mock_FixedTerm_Accounts.xlsx')
    print("Created Bank1_Mock_FixedTerm_Accounts.xlsx")

def create_accounts_loan():
    """Create sample loan accounts"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Loans"
    
    # Headers
    headers = ['LoanID', 'AccountType', 'LastPaymentDate', 'InterestFromArrearsBalance']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    data = [
        ['LOAN-001', 'Personal Loan', '2023-08-15', 2500.00],  # Overdue, arrears
        ['LOAN-002', 'Mortgage', '2023-10-01', 5000.00],       # Overdue, arrears
        ['LOAN-003', 'Business Loan', '2024-12-01', 0.00],
        ['LOAN-004', 'Auto Loan', '2023-07-20', 1200.00]      # Overdue, arrears
    ]
    
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save('Bank1_Mock_Loan_Accounts.xlsx')
    print("Created Bank1_Mock_Loan_Accounts.xlsx")

def main():
    print("Creating sample Excel files...")
    create_customer_data()
    create_accounts_cursav()
    create_accounts_fixed()
    create_accounts_loan()
    print("All sample files created successfully!")

if __name__ == '__main__':
    main()
